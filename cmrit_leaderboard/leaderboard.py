import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from cmrit_leaderboard.config import Config, LEADERBOARD_REPORT_FILE
from cmrit_leaderboard.database import Database

class Leaderboard:
    HALL_TICKET_NO = 0
    CODECHEF_USERNAME = 1
    CODECHEF_RATING = 2
    CODEFORCES_USERNAME = 3
    CODEFORCES_RATING = 4
    GEEKSFORGEEKS_USERNAME = 5
    GEEKSFORGEEKS_WEEKLY_RATING = 6
    GEEKSFORGEEKS_PRACTICE_RATING = 7
    LEETCODE_USERNAME = 8
    LEETCODE_RATING = 9
    HACKERRANK_USERNAME = 10
    HACKERRANK_RATING = 11
    CODECHEF_STATUS = 12
    CODEFORCES_STATUS = 13
    GEEKSFORGEEKS_STATUS = 14
    LEETCODE_STATUS = 15
    HACKERRANK_STATUS = 16
    TOTAL_RATING = 17
    PERCENTILE = 18

    def __init__(self):
        self.db = Database()

    def build_leaderboard(self):
        print("Fetching all users...")
        users = self.get_users()
        data = self.prepare_data(users)

        df = self.create_dataframe(data)
        df.to_excel(Config.USERS_COLLECTION + '.xlsx', index=False, engine='openpyxl')

        with pd.ExcelWriter(Config.USERS_COLLECTION + '.xlsx', engine='openpyxl', mode='a') as writer:
            self.style_worksheet(writer, df)

        print(f"Saved leaderboard report to {Config.USERS_COLLECTION + '.xlsx'}")

    def get_users(self):
        print("Fetching all users...")
        return self.db.get_all_users()

    def prepare_data(self, users):
        data = []
        for user in users:
            print(f"Processing user {user['hallTicketNo']}")
            data.append({
                'Hall Ticket No': user['hallTicketNo'],
                'CodeChef Username': user.get('codechefUsername'),
                'CodeChef Rating': user.get('codechefRating'),
                'Codeforces Username': user.get('codeforcesUsername'),
                'Codeforces Rating': user.get('codeforcesRating'),
                'GeeksforGeeks Username': user.get('geeksforgeeksUsername'),
                'GeeksforGeeks Weekly Rating': user.get('geeksforgeeksWeeklyRating'),
                'GeeksforGeeks Practice Rating': user.get('geeksforgeeksPracticeRating'),
                'Leetcode Username': user.get('leetcodeUsername'),
                'Leetcode Rating': user.get('leetcodeRating'),
                'Hackerrank Username': user.get('hackerrankUsername'),
                'Hackerrank Rating': user.get('hackerrankRating'),
                'Codechef Status': user.get('codechefStatus'),
                'Codeforces Status': user.get('codeforcesStatus'),
                'GeeksforGeeks Status': user.get('geeksforgeeksStatus'),
                'Leetcode Status': user.get('leetcodeStatus'),
                'Hackerrank Status': user.get('hackerrankStatus'),
                'Total Rating': user.get('TotalRating'),
                'Percentile': user.get('Percentile'),
            })
        return data

    def create_dataframe(self, data):
        df = pd.DataFrame(data)
        return df.sort_values(by='Percentile', ascending=False)

    def style_worksheet(self, writer, df):
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        self.apply_header_styles(worksheet)
        self.set_column_widths(worksheet, df)
        self.apply_conditional_formatting(worksheet)

        # Hide Status columns
        for i in range(self.CODECHEF_STATUS, self.TOTAL_RATING):
            worksheet.column_dimensions[get_column_letter(i + 1)].hidden = True

    def apply_header_styles(self, worksheet):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def set_column_widths(self, worksheet, df):
        for i, col in enumerate(df.columns, 1):
            # Calculate max width of header and column content
            max_length = max(df[col].astype(str).str.len().max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(i)].width = max_length

    def apply_conditional_formatting(self, worksheet):
        bad_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for row in worksheet.iter_rows(min_row=2):  # Skip header row
            for status_column in [self.CODECHEF_STATUS, self.CODEFORCES_STATUS,
                                  self.GEEKSFORGEEKS_STATUS, self.LEETCODE_STATUS,
                                  self.HACKERRANK_STATUS]:
                if row[status_column].value is False:
                    start_cell = row[status_column - 11]  # Adjust the index based on status position
                    for cell in row[start_cell.column - 1:start_cell.column + 1]:  # Adjust the range as needed
                        cell.fill = bad_fill
                        cell.border = border
                        cell.font = Font(color="9C0006")
